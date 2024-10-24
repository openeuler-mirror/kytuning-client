import logging
import json
import datetime
import openpyxl
import re
from copy import copy
from openpyxl import Workbook
from openpyxl.utils import get_column_letter ,column_index_from_string
from exportexcel import BenchMark
from logger   import log_init

current_date = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")

"""
说明:
1:程序运行过程
  main:  a.加载数据(load_data)->环境数据(__load_envinfo)
                             ->测试结果数据(__load_results)
         b.生成excel(generate_xlsx)-> export_envinfo_to_xlsx(self, sheet)
                                  ->export_results_to_xlsx(self, des_wb)
2. 生成环境数据的worksheet调用了x
3. 生成结果数据的worksheet工作机制如下:
    遍历 self.results (此为链表保存了json数据中结果部分的所有pair, )
    a.在遍历过程中, 会为每个测试工具查询是否存在该测试工具的数据, 如果查询到某个测试国工具的数据 将self_tool_sign_list[tool_name]标记为true
    ,并调用相应测试工具的worksheet生成的接口, 注意!!!该接口会收集并生成完整的worksheet,此后遍历若遍历到标记为true 的测试工具, 
    会直接跳过
    b.在遍历过程中 , 将首次查询到的某个测试工具 会调用测试工具的worksheet生成的接口, 并标记为true , 表示无需再调用该测试工具接口
3.json_file_path report_folder缺省路径需修改
"""


class ExporttoExcel():
    def __init__(self, json_file_path=r'指定缺省值', report_folder=r'指定缺省值'):
        self.if_need_xlsx = False
        self.tool_name = []  #unixbench需要合并单线程多线程数据
        self.benchmark = BenchMark()  #benchmark 依赖了exportexcel内的对象 ,
        self.report_path = report_folder + rf"\kytuning{current_date}.xlsx"  #excel文件导出的路径
        self.json_path = json_file_path  #json文件路径
        self.results = {}  #用来存放测试结果的数据 , 数组成员为dict 或者 其他
        self.envinfo = \
            {'envinfo': {
                'hwinfo': None,
                'swinfo': None,
                'nwinfo': None}
            }
        self.tools_sign_list = {'unixbench': False, 'stream': False, 'iozone': False, 'specjvm2008': False,
                                'speccpu2006': False,
                                'speccpu2017': False, 'netperf': False, 'lmbench': False, 'fio': False}
        self.tools_generate_xlsx_list = {'unixbench': Unixbench,
                                         'stream': Stream, 'iozone': Iozone, 'specjvm2008': Specjvm2008,
                                         'speccpu2006': Speccpu2006,
                                         'speccpu2017': Speccpu2017, 'netperf': Netperf, 'lmbench': Lmbench, 'fio': Fio}

    def __load_envinfo(self):
        #加载环境数据,并进行临时存储至self.envinfo
        with open(self.json_path, "r", encoding="utf-8") as f:
            json_data = json.load(f)
        if json_data['envinfo'] is None:  # !!!报错处理!!!!
            logging.error("envinfo is Missing or None in the JSON data")
            raise ValueError("envinfo is missing or None in the JSON data")
        self.envinfo['hwinfo'] = json_data['envinfo']['hwinfo']
        self.envinfo['swinfo'] = json_data['envinfo']['swinfo']
        self.envinfo['nwinfo'] = json_data['envinfo']['nwinfo']

    def __load_results(self):
        #在加载环境数据时已经对文件 json数据进行了异常处理 此处可以省略
        with open(self.json_path, "r", encoding="utf-8") as f:
            json_data = json.load(f)
            del json_data['envinfo']  #删除环境数据
            for i in json_data:
                self.results.update({i: json_data[i]})  #收集环境数据

    def export_envinfo_to_xlsx(self, sheet):
        #将环境信息导入至Excel表格中
        self.benchmark.env_dict_to_excel(sheet, self.envinfo)
        logging.info("export environmental information to excel Successfully!")
        pass

    def export_results_to_xlsx(self, des_wb):
        for key in self.results.keys():
            if (type(self.results[key]) != dict):
                continue
            for tool in self.tools_sign_list:
                if self.results[key]['tool_name'].lower() == tool.lower():
                    if (self.tools_sign_list[tool] == False):  #未被标记
                        print(f"Generating excel data for {tool}")
                        tool_generate_xlsx = self.tools_generate_xlsx_list[tool](results_dict=self.results,
                                                                                 des_wb=des_wb, key=key)
                        tool_generate_xlsx.main()  # 进入生成worksheet接口
                        logging.info(f"Excel data generation of {tool} completed")  #从接口返回,相应测试工具的excel表格已经生成完毕
                        self.tools_sign_list[tool] = True  #标记为true ,此后查询到被标记为true的测试工具 可以直接跳过 else: continue
                else:
                    continue

    def load_data(self):
        #加载两类数据的合并的函数
        self.__load_envinfo()
        self.__load_results()

    def generate_xlsx(self):
        #生成环境信息和结果数据的Excel
        wb = Workbook()
        env_ws = wb.active
        env_ws.title = 'Performance testing environment'

        self.export_envinfo_to_xlsx(env_ws)
        self.export_results_to_xlsx(wb)
        wb.save(self.report_path)
        wb.close()
        logging.info(f"File saved {self.report_path}")

    def main(self):
        # demmand = input("if need export excel(y/Y):")
        # if (demmand != 'y' and demmand != 'Y'):
        #     return

        log_init()  #初始化日志
        #try:
        self.load_data()
        self.generate_xlsx()
        # except FileNotFoundError:
        #     logging.error(f"Error: File path does not exist {self.report_path}or{self.json_path}")
        #     sys.exit(1)  # 退出程序，并返回状态码1表示错误
        # except json.JSONDecodeError:
        #     logging.error("Error: Failed to decode JSON.")
        #     sys.exit(1)
        # except Exception as e:
        #     logging.error(f"An unexpected error occurred: {e}")
        #     sys.exit(1)


"""
1.参数说明:
 ExcelBenchmark基类中:
   变量:
    a.data_list: 为二维数组:用于存放纯数据,方便直接插入到cell中 部分测试工具(unixbench,iozone,specjvm2008)会根据该变量长度计算迭代次数
    b.results_dict : 基本上就是ExporttoExcel.results 因为有且仅将该变量传入
    c.start_index : 在results_dict中首份该测试工具数据的索引
   函数:
    a.def cp_excel_format(self, des_wb, des_ws): 所有测试工具通用复制测试工具Excel表格样式
    b.def insert_command_into_excel(self, des_ws):所有测试工具通用 插入命令
    
2. speccpu2006 2017 json数据较为复杂, excel表格格式还请查看是否符合要求
3. 部分测试工具(speccpu2006 2017 ,Fio )没有根据测试的迭代次数来按列书写, 而是根据测试不同项目的项目迭代次数来按列书写
4. def cp_excel_format(self, des_wb, des_ws):
    #可以拷贝各个测试工具的Excel格式 , 注意请修改路径
    src_wb = openpyxl.load_workbook(r'C:\\32939\Desktop\gitee_repo\kytuning-client\excel_base\kytuning.xlsx')--->修改
"""


class ExcelBenchmark():
    def __init__(self, results_dict, des_wb, key):
        self.tool_name = None
        self.data_list = []  #存放纯数据(int)
        self.results_dict = results_dict  #会通过ExporttoExcel.generate_xlsx传入
        self.default_width = 14.36
        self.des_wb = des_wb
        self.key = key
        self.last_merge_col = 'B'

    def cp_excel_format(self, des_wb, des_ws):
        #可以拷贝各个测试工具的Excel格式 , 注意请修改路径
        src_wb = openpyxl.load_workbook(r'指定excel_base/kytuning.xlsx的路径')
        src_ws = src_wb[self.tool_name]

        # 复制列宽
        for column in range(1, src_ws.max_column + 1):
            col_letter = get_column_letter(column)
            if col_letter in src_ws.column_dimensions:
                des_ws.column_dimensions[col_letter].width = src_ws.column_dimensions[col_letter].width
            else:
                des_ws.column_dimensions[col_letter].width = self.default_width
        # 复制行高
        for row_num, row_dim in src_ws.row_dimensions.items():
            des_ws.row_dimensions[row_num].height = row_dim.height
        # 复制单元格内容和样式
        for row in src_ws.iter_rows():
            for cell in row:
                new_cell = des_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
        # 合并指定的单元格
        for row_num in range(1, 4):
            des_ws.merge_cells(f"A{row_num}:{self.last_merge_col}{row_num}")
        logging.info(f"Workbook format of {self.tool_name} copied Successfully")

    def insert_command_into_excel(self, des_ws):
        command = self.results_dict[self.key].get('execute_cmd')
        if (command != None):
            col = column_index_from_string(self.last_merge_col)+1
            des_ws.cell(row=2, column=col, value=command)
        else:
            logging.info("the command is empty")


class Unixbench(ExcelBenchmark):
    def __init__(self, results_dict, des_wb, key):
        super().__init__(results_dict, des_wb, key)
        self.tool_name = r'Unixbench'
        self.iteration = 0  #由于之前测试出现过仅生成单线程数据的情况,迭代次数根据单线程数据列表个数来决定

    def main(self):
        self.collect_data()
        self.export_data_to_xlsx()

    def collect_data(self):  #数据收集后进行删除
        #找到目标数据的开头索引
        for key in self.results_dict:
            if (type(self.results_dict[key]) != dict):
                continue
            if (self.results_dict[key]['tool_name'].lower() != self.tool_name.lower()):
                continue
            #经过两轮排除后的字典一定符合要求 需要将单线程多线程数据写入
            #创建临时数组来暂存数据
            temp = []
            if ('单线程' in self.results_dict[key]):
                """根据实际情况 ,有时不会生成多线程数据 故采用检测单线程数据来确定迭代次数"""
                self.iteration += 1
                for val in self.results_dict[key]['单线程'].values():
                    temp.append(val)
                self.data_list.append(temp)
            else:
                for val in self.results_dict[key]['多线程'].values():
                    temp.append(val)
                self.data_list.append(temp)
            logging.info("Data collection of Unixbench completed")

    def export_data_to_xlsx(self):
        des_ws = self.des_wb.create_sheet(title=self.tool_name)
        self.cp_excel_format(self.des_wb, des_ws)
        self.insert_command_into_excel(des_ws)
        for i in range(self.iteration):
            des_ws.cell(row=1, column=i + 3, value=f"test#{i + 1}")
            for idx, value in enumerate(self.data_list[i], start=4):  #单线程数据,在表格中 第4行开始记录测试数据
                des_ws.cell(row=idx, column=i + 3, value=value)
            if (i + self.iteration < len(self.data_list)):  #判定是否存在多线程数据
                for idx, value in enumerate(self.data_list[self.iteration + i], start=17):  #多线程数据 , 在表格中第17行开始记录
                    des_ws.cell(row=idx, column=i + 3, value=value)
            else:
                logging.warning("Missing multi-threaded data of Unixbench")


class Stream(ExcelBenchmark):
    def __init__(self, results_dict, des_wb, key):
        super().__init__(results_dict, des_wb, key)  # 正确调用父类构造函数
        self.tool_name = 'Stream'

    def main(self):
        self.collect_data()
        self.export_data_to_xlsx()

    def collect_data(self):
        for key in self.results_dict:
            if (type(self.results_dict[key]) != dict):
                continue
            if (self.results_dict[key]['tool_name'].lower() != self.tool_name.lower()):
                continue
            temp = []
            if ('单线程' in self.results_dict[key]):
                for val in self.results_dict[key]['单线程'].values():
                    temp.append(val)
                self.data_list.append(temp)
            elif ('多线程' in self.results_dict[key]):
                for val in self.results_dict[key]['多线程'].values():
                    temp.append(val)
                self.data_list.append(temp)
        logging.info("Data collection of Stream completed")

    def export_data_to_xlsx(self):
        des_ws = self.des_wb.create_sheet(title=self.tool_name)
        self.cp_excel_format(self.des_wb, des_ws)
        self.insert_command_into_excel(des_ws)
        iteration = int(len(self.data_list) / 2)
        for i in range(iteration):
            des_ws.cell(row=1, column=i + 3, value=f"test#{i + 1}")
            for idx, value in enumerate(self.data_list[2 * i], start=4):  #单线程数据 从第4行开始书写
                des_ws.cell(row=idx, column=i + 3, value=value)
            for idx, value in enumerate(self.data_list[2 * i + 1], start=9):  #多线程数据 从第9行开始书写
                des_ws.cell(row=idx, column=i + 3, value=value)


class Iozone(ExcelBenchmark):
    def __init__(self, results_dict, des_wb, key):
        super().__init__(results_dict, des_wb, key)
        self.tool_name = 'Iozone'
        self.iteration = 0

    def main(self):
        self.collect_data()
        self.export_data_to_xlsx()
        pass

    def collect_data(self):
        for key in self.results_dict:
            if (type(self.results_dict[key]) != dict):
                continue
            if (self.results_dict[key]['tool_name'].lower() != self.tool_name.lower()):
                continue
            temp = []
            if ('测试记录' in self.results_dict[key]):
                for val in self.results_dict[key]['测试记录'].values():
                    temp.append(val)
                self.data_list.append(temp)
        logging.info("Data collection of Iozone completed")

    def export_data_to_xlsx(self):
        des_ws = self.des_wb.create_sheet(title=self.tool_name)
        self.cp_excel_format(self.des_wb, des_ws)
        self.insert_command_into_excel(des_ws)
        self.iteration = int(len(self.data_list) / 2)
        for i in range(self.iteration):
            des_ws.cell(row=1, column=i + 3, value=f"test#{i + 1}")
            for idx, value in enumerate(self.data_list[i], start=4):  #测试记录 从4行开始书写
                des_ws.cell(row=idx, column=i + 3, value=value)


class Lmbench(ExcelBenchmark):
    def __init__(self, results_dict, des_wb, key):
        super().__init__(results_dict, des_wb, key)  # 正确调用父类构造函数
        self.tool_name = 'Lmbench'

    def main(self):
        self.collect_data()
        self.export_data_to_xlsx()

    def collect_data(self):
        for key in self.results_dict:
            if (type(self.results_dict[key]) != dict):
                continue
            if (self.results_dict[key]['tool_name'].lower() != self.tool_name.lower()):
                continue
            if ('items' in self.results_dict[key]):
                for outer_dict in self.results_dict[key]['items']:
                    temp = []
                    for inner_dict in outer_dict:
                        for pairs in inner_dict.values():
                            for data in pairs.values():
                                temp.append(data)
                    self.data_list.append(temp)
        logging.info("Data collection of Lmbench completed")

    def export_data_to_xlsx(self):
        if self.tool_name in self.des_wb.sheetnames:
            des_ws = self.des_wb[self.tool_name]
        else:
            des_ws = self.des_wb.create_sheet(title=self.tool_name)
        self.cp_excel_format(self.des_wb, des_ws)
        self.insert_command_into_excel(des_ws)
        iteration = len(self.data_list)
        for i in range(iteration):
            des_ws.cell(row=1, column=i + 3, value=f"test#{i + 1}")
            for idx, value in enumerate(self.data_list[i], start=4):  #测试记录
                des_ws.cell(row=idx, column=i + 3, value=value)


class Fio(ExcelBenchmark):
    def __init__(self, results_dict, des_wb, key):
        super().__init__(results_dict, des_wb, key)  # 正确调用父类构造函数
        self.tool_name = 'Fio'
        self.data_list = {'4K': {'read':[] , 'write':[],'randread':[],'randwrite':[]},
                          '16K':{'read':[] , 'write':[]},
                          '64K':{'read':[] , 'write':[]},
                          '128K':{'read':[] , 'write':[]},
                          '1M':{'read':[] , 'write':[]}}
        self.data_insert_idx = {'4K': {'read':4 , 'write':8,'randread':12,'randwrite':16},
                          '16K':{'read':20 , 'write':24},
                          '64K':{'read':28 , 'write':32},
                          '128K':{'read':36 , 'write':40},
                          '1M':{'read':44 , 'write':48}}
    def main(self):
        self.collect_data()
        self.export_data_to_xlsx()

    def collect_data(self):
        for key in self.results_dict:
            if (type(self.results_dict[key]) != dict):
                continue
            if (self.results_dict[key]['tool_name'].lower() != self.tool_name.lower()):
                continue
            match = re.search(r'(\d+[KM])-(read|write|randread|randwrite)', key)
            if not match: continue
            temp = []
            block_size = match.group(1)
            operation = match.group(2)
            for data in self.results_dict[key]['items'].values():
                temp.append(data)
            self.data_list[block_size][operation].append(temp)
        return

    def export_data_to_xlsx(self):
        des_ws = self.des_wb.create_sheet(title=self.tool_name)
        self.cp_excel_format(self.des_wb, des_ws)
        self.insert_command_into_excel(des_ws)
        # 因为生成的excel表格无法通过迭代次数确定列, 因此引入column_idx
        for block_size  in self.data_list:
            for operation in self.data_list[block_size]:
                for list_idx in range(len(self.data_list[block_size][operation])):
                    des_ws.cell(row=1, column=list_idx + 3, value=f"test#{list_idx + 1}")
                    for idx, value in enumerate(self.data_list[block_size][operation][list_idx], start=self.data_insert_idx[block_size][operation]):
                        des_ws.cell(row=idx, column= list_idx+3, value=value)


class Specjvm2008(ExcelBenchmark):
    def __init__(self, results_dict, des_wb, key):
        super().__init__(results_dict, des_wb, key)
        self.key = key
        self.tool_name = 'Specjvm2008'
        self.iteration = 0

    def main(self):
        self.collect_data()
        self.export_data_to_xlsx()
        pass

    def collect_data(self):
        for key in self.results_dict:
            if (type(self.results_dict[key]) != dict):
                continue
            if (self.results_dict[key]['tool_name'].lower() != self.tool_name.lower()):
                continue
            temp = []
            if ('items' in self.results_dict[key]):
                if 'base' in self.results_dict[key]['items']:
                    for data in self.results_dict[key]['items']['base'].values():
                        temp.append(data)
                elif 'peak' in self.results_dict[key]['items']:
                    for data in self.results_dict[key]['items']['peak'].values():
                        temp.append(data)
                self.data_list.append(temp)
        logging.info("Data collection of Specjvm2008 completed")

    def export_data_to_xlsx(self):
        des_ws = self.des_wb.create_sheet(title=self.tool_name)
        self.cp_excel_format(self.des_wb, des_ws)
        self.insert_command_into_excel(des_ws)
        iteration = int(len(self.data_list) / 2)  # 默认base 和peak 的数据会一起生成
        for i in range(iteration):
            des_ws.cell(row=1, column=i + 3, value=f"test#{i + 1}")
            for idx, value in enumerate(self.data_list[i], start=4):  #base
                des_ws.cell(row=idx, column=i + 3, value=value)
            for idx, value in enumerate(self.data_list[self.iteration + i], start=16):  #peak
                des_ws.cell(row=idx, column=i + 3, value=value)


class Speccpu2006(ExcelBenchmark):
    def __init__(self, results_dict, des_wb, key):
        super().__init__(results_dict, des_wb, key)
        self.tool_name = 'Speccpu2006'
        self.base_datas = []  #单线程 int数据
        self.peak_datas = []  #多线程  int数据
        self.length_of_data = 62
        self.last_merge_col ='C'

    def main(self):
        self.collect_data()
        self.export_data_to_xlsx()

    def collect_data(self):
        pattern = r'cpu2006-\d+\.\d+-(base)-'
        base_temp = []
        peak_temp = []
        for key in self.results_dict:
            if (type(self.results_dict[key]) != dict):
                continue
            if (self.results_dict[key]['tool_name'].lower() != self.tool_name.lower()):
                continue
            if (re.search(pattern, key)):  #base 数据
                for outer_pair in self.results_dict[key]['items'].values():
                    for inner_key in outer_pair:
                        if (inner_key == 'base'):
                            for data in outer_pair['base'].values():
                                base_temp.append(data)
            if (len(base_temp) == self.length_of_data):
                self.base_datas.append(base_temp)
                base_temp = []

            else:  #peak数据
                for outer_pair in self.results_dict[key]['items'].values():
                    for inner_key in outer_pair:
                        if (inner_key == 'peak'):
                            for data in outer_pair['peak'].values():
                                peak_temp.append(data)
            if (len(peak_temp) == self.length_of_data):
                self.peak_datas.append(peak_temp)
                peak_temp = []

    def export_data_to_xlsx(self):
        des_ws = self.des_wb.create_sheet(title=self.tool_name)
        self.cp_excel_format(self.des_wb, des_ws)
        self.insert_command_into_excel(des_ws)
        column_idx = 3+1
        for i in range(len(self.base_datas)):
            des_ws.cell(row=1, column=column_idx, value=f"base#{i + 1}")
            for idx, value in enumerate(self.base_datas[i], start=4):
                des_ws.cell(row=idx, column=column_idx, value=value)
            column_idx += 1
        for i in range(len(self.peak_datas)):
            des_ws.cell(row=1, column=column_idx, value=f"base#{i + 1}")
            for idx, value in enumerate(self.base_datas[i], start=4):
                des_ws.cell(row=idx, column=column_idx, value=value)
            column_idx += 1


class Speccpu2017(ExcelBenchmark):
    def __init__(self, results_dict, des_wb, key):
        super().__init__(results_dict, des_wb, key)
        self.tool_name = 'Speccpu2017'
        self.if_data_exist = False
        self.pattern =''  #rate or speed
        self.last_merge_col = 'C'
        self.data_list = {'base': {'single': [], 'multi': []},
                          'peak': {'single': [], 'multi': []}
                          }
        self.data_insert_idx = { }
        self.length_of_data = -1

    def collect_data(self):
        temp = []
        for key in self.results_dict:
            if (type(self.results_dict[key]) != dict):
                continue
            if not (self.pattern in key):
                continue
            self.if_data_exist = True
            single_or_multi = ''
            peak_or_base = ''

            if 'single' in key or 'multi' in key:
                single_or_multi = 'single' if 'single' in key else 'multi'

            if re.search(r'\b(peak|base)\b', key):
                peak_or_base = re.search(r'\b(peak|base)\b', key).group() or 'base'

            for  outer_pair in self.results_dict[key]['items'].values():
                for data in outer_pair[peak_or_base].values():
                    temp.append(data)
            if(len(temp) == self.length_of_data):
                self.data_list[peak_or_base][single_or_multi].append(temp)
                temp = []


    def export_data_to_xlsx(self):
        if not(self.if_data_exist): return
        des_ws = self.des_wb.create_sheet(title=self.tool_name)
        self.cp_excel_format(self.des_wb, des_ws)
        self.insert_command_into_excel(des_ws)
        column_idx = 3+1
        for mode in self.data_list:
            if(self.data_list[mode]['single']!=[]):
                for i in range(len(self.data_list[mode]['single'])):
                    des_ws.cell(row=1, column=column_idx, value=f"{mode}#{i + 1}")
                    for idx, value in enumerate(self.data_list[mode]['single'][i], start=self.data_insert_idx['single']):
                        des_ws.cell(row=idx, column=column_idx, value=value)
            if(self.data_list[mode]['multi']!= []) :
                for i in range(len(self.data_list[mode]['multi'])):
                    des_ws.cell(row=1, column=column_idx ,value=f"{mode}#{i + 1}")
                    for idx, value in enumerate(self.data_list[mode]['multi'][i], start= self.data_insert_idx['multi']):
                        des_ws.cell(row=idx, column=column_idx, value=value)
                    column_idx += 1

    def main(self):
        Interface = Speccpu2017_rate(self.results_dict , self.des_wb , self.key)
        Interface.main()
        Interface = Speccpu2017_speed(self.results_dict , self.des_wb ,self.key)
        Interface.main()


class Speccpu2017_rate(Speccpu2017):
    def __init__(self, results_dict, des_wb, key):
        super().__init__(results_dict, des_wb, key)
        self.tool_name = 'Speccpu2017(rate)'
        self.data_list = {'base': {'single':[],'multi':[]},
                          'peak':{'single':[],'multi':[]}
                          }
        self.data_insert_idx = {'single':4  , 'multi': 29}
        self.length_of_data = 25
        self.pattern = 'rate'

    def main(self):
        self.collect_data()
        self.export_data_to_xlsx()


class Speccpu2017_speed(Speccpu2017):
    def __init__(self, results_dict, des_wb, key):
        super().__init__(results_dict, des_wb, key)
        self.tool_name = 'Speccpu2017(speed)'
        self.data_list = {'base': {'single': [], 'multi': []},
                          'peak': {'single': [], 'multi': []}
                          }
        self.data_insert_idx = {'single': 4, 'multi':26 }
        self.length_of_data = 22
        self.pattern = 'speed'

    def main(self):
        self.collect_data()
        self.export_data_to_xlsx()


class Netperf(ExcelBenchmark):
    def __init__(self, results_dict, start_index, des_wb):
        super().__init__(results_dict, start_index, des_wb)
        self.tool_name = 'Netperf'

    def main(self):
        self.collect_data()
        self.export_data_to_xlsx()

    def collect_data(self):
        for key in self.results_dict:
            if (type(self.results_dict[key]) != dict):
                continue
            if (self.results_dict[key]['tool_name'].lower() != self.tool_name.lower()):
                continue
            temp = []
            if ('items' in self.results_dict[key]):
                for data in self.results_dict[key]['items'].values():
                    temp.append(data)
            self.data_list.append(temp)

    def export_data_to_xlsx(self):
        if self.tool_name in self.des_wb.sheetnames:
            des_ws = self.des_wb[self.tool_name]
        else:
            des_ws = self.des_wb.create_sheet(title=self.tool_name)
        self.cp_excel_format(self.des_wb, des_ws)
        self.insert_command_into_excel(des_ws)
        iteration = int(len(self.data_list) / 2)
        for i in range(iteration):
            des_ws.cell(row=1, column=i + 3, value=f"test {i + 1}")
            for idx, value in enumerate(self.data_list[i], start=4):  # 测试记录
                des_ws.cell(row=idx, column=i + 3, value=value)


if __name__ == '__main__':
    # json_file_path = input("please input the json file path:")
    # report_folder = input("please input the report directory:")
    # if (json_file_path and report_folder):
    #     test = ExporttoExcel(json_file_path, report_folder)
    # else:
    test = ExporttoExcel()
    test.main()
