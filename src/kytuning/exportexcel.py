"""
 * Copyright (c) KylinSoft  Co., Ltd. 2024.All rights reserved.
 * PilotGo-plugin licensed under the Mulan Permissive Software License, Version 2. 
 * See LICENSE file for more details.
 * Author: wangqingzheng <wangqingzheng@kylinos.cn>
 * Date: Thu Dec 14 10:56:24 2023 +0800
"""
#!/usr/bin/python3
# 功能：从环境/测试结果json文件导出到excel
# 依赖库：openpyxl

__all__ = ['ExportXlsx']
import os
import re
import json
import difflib
import argparse
from base64 import b64decode
from openpyxl import Workbook, load_workbook, utils
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment
from openpyxl.styles import Font, PatternFill, Side, Border

ROOT_PATH = os.path.abspath(os.path.dirname(__file__))

TOOL_NAME_UB = "Unixbench"
TOOL_NAME_CPU06 = "Speccpu2006"
TOOL_NAME_CPU17 = "Speccpu2017"
TOOL_NAME_IOZONE = "Iozone"
TOOL_NAME_JVM08 = "Specjvm2008"
TOOL_NAME_LMB = "Lmbench"
TOOL_NAME_STREAM = "Stream"
TOOL_NAME_NETPERF = "Netperf"
TOOL_NAME_FIO = "Fio"


class BenchMark(object):
    def __init__(self):
        super().__init__()
        self.items = None
        self.tool_name = None
        self.tool_version = None  # 暂时未使用

        self.sheet_env_title = "性能测试环境"
        self.index_start = 1
        self.text_cmd = "执行命令:"
        self.text_modify_args = "修改参数:"
        self.value_cmd = "测试None"
        self.value_modify_args = "测试None"

        self.alignment_center = Alignment(
            horizontal='center', vertical='center', wrapText=True)
        self.alignment_left = Alignment(
            horizontal='left', vertical='center', wrapText=True)
        self.ret_col_1_width = 30
        self.ret_col_2_width = 20
        self.ret_col_data_width = 15

        self.ret_row_1_height = 50

        # 起始行列
        self.row_point_start = 1
        self.col_point_start = 1

        # 初始化行与列的高度和宽度,可在各个子类中覆盖.
        self.cols_width = [self.ret_col_1_width]
        self.rows_height = [
            25,                    # 开始行
            25,                    # 命令行
            self.ret_row_1_height] # 参数行

        # excel_style
        thin = Side(border_style="thin", color="000000")
        self.border_thin = Border(
            top=thin, left=thin, right=thin, bottom=thin)
        self.color_title = PatternFill("solid", fgColor="CC99FF")
        self.font_title = Font(name="宋体", size=20, bold=True, color="993300")

        self.color_cmd = PatternFill("solid", fgColor="CCFFFF")
        self.font_cmd = Font(name="宋体", size=13, bold=True, color="000000")

        self.color_col_top = PatternFill("solid", fgColor="99CCFF")
        self.font_col_top = Font(name="宋体", size=18, color="000000")

        self.color_item_1 = PatternFill("solid", fgColor="FFCC99")
        self.font_item_1 = Font(name="宋体", size=13, color="000000")
        self.color_item_2 = PatternFill("solid", fgColor="CCFFCC")
        self.font_item_2 = Font(name="宋体", size=11, color="000000")

        self.color_data = PatternFill()
        self.font_data = Font(name="宋体")

    @property
    def row_data_start(self):
        return self.row_point_start + len(self.rows_height)

    @property
    def col_data_start(self):
        return self.col_point_start + len(self.cols_width)

    def set_cell_style(
        self, sheet, row_idx, col_idx, value,
            alg, fill=PatternFill(), font=Font(name="宋体")):
        cell = sheet.cell(row_idx, col_idx, value)
        cell.border = self.border_thin
        cell.alignment = alg
        cell.fill = fill
        cell.font = font

    def merge_col_cell(
            self, worksheet: Worksheet, N: str, row_A: int, row_B: int):
        if row_B - row_A > 0:
            alignment = Alignment(horizontal='center', vertical='center')
            merge_row = str(N) + str(row_A) + ':' + str(N) + str(row_B)
            worksheet.merge_cells(merge_row)
            merge_row = str(N)+str(row_A)
            worksheet[merge_row].alignment = alignment

    def unmerge_col_cell(
            self, worksheet: Worksheet, N: str, row_A: int, row_B: int):
        if row_B - row_A > 0:
            merge_row = str(N) + str(row_A) + ':' + str(N) + str(row_B)
            worksheet.unmerge_cells(merge_row)

    def merge_col_cell_by_value(self, sheet: Worksheet, N: str, row_A: int, row_B: int, value = None):
        if row_B - row_A > 0:
            merge_row = 0
            if row_A > self.row_point_start and value is not None:
                for i in range(self.row_point_start, row_A):
                    cval = sheet.cell(row_A - i, utils.column_index_from_string(N)).value
                    if cval is None:
                        continue
                    if cval == value:
                        merge_row = i
                    break
                if merge_row > 0:
                    self.unmerge_col_cell(sheet, N, row_A - merge_row, row_A - 1)
            self.merge_col_cell(sheet, N, row_A - merge_row, row_B)

    def merge_row_cell(
            self, worksheet: Worksheet, row_N: int, col_A: int, col_B: int):
        if col_B - col_A > 0:
            col_A = utils.get_column_letter(col_A)
            col_B = utils.get_column_letter(col_B)
            alignment = Alignment(horizontal='center', vertical='center')
            merge_row = str(col_A) + str(row_N) + ':' + str(col_B) + str(row_N)
            worksheet.merge_cells(merge_row)
            merge_row = str(col_A)+str(row_N)
            worksheet[merge_row].alignment = alignment

    def max_column_by_rows(self, sheet: Worksheet, min_row = None, max_row = None):
        if min_row is None and max_row is None:
            return sheet.max_column

        rows_vcount = []
        min_row = min_row or self.row_point_start
        max_row = max_row or sheet.max_row
        for row in sheet.iter_rows(min_row = min_row, max_row = max_row, values_only = True):
            row_vcount = len(row)
            for idx in range(row_vcount, self.col_point_start - 1, -1):
                if row[idx - 1] is None:
                    row_vcount -= 1
                else:
                    break
            rows_vcount.append(row_vcount)
        return max(tuple(rows_vcount)) if len(rows_vcount) > 0 else sheet.max_column

    def find_row_point(self, sheet: Worksheet, data: dict):
        if sheet.max_row >= self.row_data_start:
            col_values = [None for i in range(0, len(data))]
            for row in sheet.iter_rows(min_row = self.row_data_start):
                for idx in range(0, len(data)):
                    if row[self.col_point_start - 1 + idx].value is not None:
                        col_values[idx] = row[self.col_point_start - 1 + idx].value
                for idx, _v in enumerate(data.values()):
                    if col_values[idx] != _v:
                        break
                else:
                    return row[self.col_point_start - 1].row
        return sheet.max_row + 1

    def find_col_point(self, sheet: Worksheet, row: int):
        if sheet.max_column >= self.col_data_start:
            for col in sheet.iter_cols(min_col = self.col_data_start):
                if col[row - 1].value is None:
                    if col[self.row_point_start - 1 + 2].value is None or col[self.row_point_start - 1 + 2].value == self.value_modify_args:
                        return col[self.row_point_start].column
        return sheet.max_column + 1

    def set_col_items(self, sheet: Worksheet):
        point_row = self.row_point_start
        point_col = self.col_point_start
        cols_num = len(self.cols_width)
        # 开始
        self.set_cell_style(sheet, point_row, point_col, self.tool_name,
                            self.alignment_center, self.color_title, self.font_title)
        self.merge_row_cell(sheet, point_row, point_col, point_col + cols_num - 1)
        # 执行命令
        point_row += 1
        self.set_cell_style(sheet, point_row, point_col, self.text_cmd,
                            self.alignment_center, self.color_cmd, self.font_cmd)
        self.merge_row_cell(sheet, point_row, point_col, point_col + cols_num - 1)
        # 修改参数
        point_row += 1
        self.set_cell_style(sheet, point_row, point_col, self.text_modify_args,
                            self.alignment_center, self.color_cmd, self.font_cmd)
        self.merge_row_cell(sheet, point_row, point_col, point_col + cols_num - 1)

        for idx, wth in enumerate(self.cols_width):
            sheet.column_dimensions[utils.get_column_letter(idx + self.col_point_start)].width = wth

        for idx, hgh in enumerate(self.rows_height):
            sheet.row_dimensions[idx + self.row_point_start].height = hgh
        pass

    def set_col_title(self, sheet: Worksheet, col_point: int):
        if col_point > sheet.max_column:
            sheet.column_dimensions[utils.get_column_letter(col_point)].width = self.ret_col_data_width
            self.set_cell_style(sheet, self.row_point_start, col_point, self.tool_name + '#' + str(col_point - 2),
                                self.alignment_center, self.color_col_top, self.font_col_top)
            self.set_cell_style(sheet, self.row_point_start + 1, col_point, self.value_cmd,
                                self.alignment_center, self.color_data, self.font_data)
            self.set_cell_style(sheet, self.row_point_start + 2, col_point, self.value_modify_args,
                                self.alignment_center, self.color_data, self.font_data)
        else:
            if sheet.cell(self.row_point_start + 1, col_point).value != self.value_cmd:
                self.set_cell_style(sheet, self.row_point_start + 1, col_point, sheet.cell(self.row_point_start + 1, col_point).value + "\n" + self.value_cmd,
                                self.alignment_center, self.color_data, self.font_data)
        pass

    def set_row_header(self, sheet: Worksheet, row_point: int, ret_dict):
        if row_point > sheet.max_row:
            pass
            # self.set_cell_style(sheet, row_point, 1, ret_dict['rw'], self.alignment_center, self.color_item_1, self.font_item_1)
            # for i, _l in enumerate(self.items["items"]):
            #     self.set_cell_style(sheet, row_point + i, 2, _l, self.alignment_center, self.color_item_2, self.font_item_2)
            # self.merge_col_cell(sheet, "A", row_point, row_point + len(self.items["items"]) - 1)
        pass

    def ret_to_dict(self, file: str): ...

    def env_dict_to_excel(self, sheet: Worksheet, env_dict: dict):
        if not env_dict:
            return
        text_sheet_title = "性能测试环境统计表"
        _env_dict = env_dict
        index_start = 1
        self.set_cell_style(sheet, index_start, index_start,
                            text_sheet_title, self.alignment_center,
                            self.color_title, self.font_title)
        sheet.merge_cells('A1:%s1' % (
            utils.get_column_letter(3+1)))  # +1是包含字典中值的那一列
        sheet.column_dimensions["A"].width = 15
        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 15
        sheet.column_dimensions["D"].width = 50
        sheet.row_dimensions[1].height = 30

        base64_list = ["sysctl", "sysconf",
                       "systemctlinfo", "driverinfo", "rpmlist", "ipclist"]

        row_start_idx = 5
        _row_up = 0
        _row = 1
        A_names = list(_env_dict["envinfo"].keys())
        A_names_len = len(A_names)
        for A_i in range(A_names_len):     # 遍历第一列中的内容
            self.set_cell_style(
                sheet, _row+row_start_idx,
                1,  A_names[A_i], self.alignment_center,
                self.color_title, self.font_title)  # 填充第一列中的内容  'hwinfo'
            _row_A = _row
            # print(type(env_dict["envinfo"][A_names[A_i]]))
            if type(_env_dict[A_names[A_i]]) != str:
                key_names = list(_env_dict[A_names[A_i]].keys())
                keys_len = len(key_names)
                for i in range(keys_len):   # 遍历第二列中的内容
                    # print(key_names[i])
                    self.set_cell_style(
                        sheet, _row + row_start_idx,
                        2, key_names[i], self.alignment_center,
                        self.color_item_1, self.font_item_1)   # 填充第二列中的内容
                    value_type = type(
                        _env_dict[A_names[A_i]][key_names[i]])
                    if value_type == str:
                        cell_str = _env_dict[A_names[A_i]
                                                        ][key_names[i]]
                        self.set_cell_style(
                            sheet, _row+row_start_idx, 4,
                            cell_str,
                            self.alignment_left,
                            self.color_data, self.font_data)
                        _row = _row + 1
                    elif value_type == dict:
                        dict_keys_name = list(
                            _env_dict[A_names[A_i]
                                                 ][key_names[i]].keys())
                        dict_keys_len = len(dict_keys_name)
                        _row_up = _row
                        for j in range(dict_keys_len):
                            self.set_cell_style(
                                sheet, _row+row_start_idx, 3,
                                dict_keys_name[j], self.alignment_center,
                                self.color_item_2, self.font_item_2)  # 填充第三列
                            dict_keys_type = type(
                                _env_dict[A_names[A_i]
                                                     ][key_names[i]
                                                       ][dict_keys_name[j]])
                            if dict_keys_type == str:
                                cell_str = _env_dict[A_names[A_i]][key_names[i]
                                                         ][dict_keys_name[j]]
                                if any(dict_keys_name[j]
                                       in s for s in base64_list):
                                    cell_str = b64decode(
                                        cell_str).decode("ascii")
                                col_n = 4
                                while True:
                                    if len(cell_str) > 8100:
                                        _cell_str = cell_str[:8100]
                                        if not _cell_str.endswith('\n'):
                                            _line_end = _cell_str.rfind('\n')
                                            _cell_str = cell_str[:_line_end]
                                        cell_str = cell_str[len(_cell_str):]

                                        self.set_cell_style(
                                            sheet, _row+row_start_idx, col_n,
                                            _cell_str.encode("ascii"),
                                            self.alignment_left,
                                            self.color_data, self.font_data)
                                        col_n += 1
                                        sheet.column_dimensions[
                                            utils.get_column_letter(
                                                col_n)].width = 50
                                    else:
                                        self.set_cell_style(
                                            sheet, _row+row_start_idx, col_n,
                                            cell_str.encode("utf8"),
                                            self.alignment_left,
                                            self.color_data, self.font_data)
                                        break
                            _row = _row + 1
                        # 合并第二列中的单元格
                        self.merge_col_cell(
                            sheet, 'B',
                            _row_up+row_start_idx, _row-1+row_start_idx)
                    elif value_type == list:
                        _row_up = _row
                        for list_l in _env_dict[A_names[A_i]
                                                           ][key_names[i]]:
                            if type(list_l) == str:  # memType
                                # print(list_l)
                                self.set_cell_style(
                                    sheet, _row+row_start_idx, 4, list_l,
                                    self.alignment_left,
                                    self.color_data, self.font_data)
                                _row = _row + 1
                            elif type(list_l) == dict:  # disk
                                list_key_names = list(list_l.keys())
                                for list_key_name in list_key_names:
                                    # print(list_key_name)
                                    # 填充第三列中的内容
                                    self.set_cell_style(
                                        sheet, _row+row_start_idx, 3,
                                        list_key_name,
                                        self.alignment_center,
                                        self.color_item_2, self.font_item_2)
                                    if type(list_l[list_key_name]) == str:
                                        self.set_cell_style(
                                            sheet, _row + row_start_idx, 4,
                                            list_l[list_key_name],
                                            self.alignment_left,
                                            self.color_data, self.font_data)
                                    _row = _row + 1
                        # 合并第二列中的单元格
                        self.merge_col_cell(
                            sheet, 'B',
                            _row_up+row_start_idx, _row-1+row_start_idx)
                # 合并第一列中的单元格
                self.merge_col_cell(sheet, 'A',
                                    _row_A+row_start_idx, _row-1+row_start_idx)
            else:
                self.set_cell_style(
                    sheet, _row +
                    row_start_idx, 4, _env_dict[A_names[A_i]],
                    self.alignment_left,
                    self.color_data, self.font_data)

    def ret_dict_to_excel(self, workbook: Workbook, ret_dict: dict): ...


class Unixbench(BenchMark):
    def __init__(self):
        super().__init__()
        self.tool_name = TOOL_NAME_UB
        self.items = ["Dhrystone 2 using register variables(lps)",
                      "Double-Precision Whetstone(MWIPS)",
                      "Execl Throughput(lps)",
                      "File Copy 1024 bufsize 2000 maxblocks(KBps)",
                      "File Copy 256 bufsize 500 maxblocks(KBps)",
                      "File Copy 4096 bufsize 8000 maxblocks(KBps)",
                      "Pipe Throughput(lps)",
                      "Pipe-based Context Switching(lps)",
                      "Process Creation(lps)",
                      "Shell Scripts (1 concurrent)(lpm)",
                      "Shell Scripts (8 concurrent)(lpm)",
                      "System Call Overhead(lps)",
                      "Index Score(sum)"]
        self.thread = ["单线程", "多线程"]
        self.cmd = "执行命令："
        self.modify_argv = "修改参数："
        self.xlsx_col = self.thread
        self.cpu_num_flag = "in system; run"
        self.thread_flag = "1 parallel"
        self.read_items_flag = "BASELINE"
        self.cols_width = [10, self.ret_col_1_width]

    def set_row_header(self, sheet: Worksheet, row_point: int, data: dict):
        if row_point > sheet.max_row:
            self.set_cell_style(sheet, row_point, self.col_point_start, data["tune"], self.alignment_center, self.color_item_1, self.font_item_1)
            for i, _l in enumerate(self.items):
                self.set_cell_style(sheet, row_point + i, self.col_point_start + 1, _l, self.alignment_left, self.color_item_2, self.font_item_2)
            self.merge_col_cell(sheet, utils.get_column_letter(self.col_point_start), row_point, row_point + len(self.items) - 1)
        pass

    def ret_to_dict(self, file: str):
        """测试结果解析"""
        ret_dict = {"tool_name": self.tool_name}
        _thread = ''
        _max_thread = ''
        _read_flag = False
        score_list = ['' for i in self.items]
        with open(file, 'r') as f:
            file_lines = f.readlines()
        for line in file_lines:
            if self.cpu_num_flag in line:
                maxcpu = re.findall(
                    r"-?\d+\.?\d*e?-?\d*?", line)[0]
                _max_thread = maxcpu + " parallel"
            if self.thread_flag in line:
                _thread = self.thread[0]
                continue
            elif _max_thread in line:
                _thread = self.thread[1]
                continue
            if self.read_items_flag in line:
                _read_flag = True
                continue
            if _read_flag:
                _score_list = score_list
                for i, key in enumerate(self.items):
                    if key[:-7] in line:
                        temp = re.findall(r"-?\d+\.?\d*e?-?\d*?", line)
                        if temp:
                            _score_list[i] = temp[-1]
                        elif "==" not in line:
                            _score_list[i] = ''
                        if key == self.items[-1]:
                            ret_dict[_thread] = dict(
                                zip(self.items, _score_list))
                            _read_flag = False
                            break
        if len(ret_dict.keys()) > 1:
            return ret_dict
        else:
            return None

    def ret_dict_to_excel(self, workbook: Workbook, ret_dict: dict):
        sheet = None
        if self.tool_name in workbook.sheetnames:  # 已存在sheet
            sheet = workbook[self.tool_name]
        else:   # 不存在sheet
            sheet = workbook.create_sheet(self.tool_name)
            self.set_col_items(sheet)

        for tune in self.thread:
            if tune not in ret_dict.keys():
                continue
            row_point = self.find_row_point(sheet, {"tune": tune})
            self.set_row_header(sheet, row_point, {"tune": tune})

            col_point = self.find_col_point(sheet, row_point)
            self.set_col_title(sheet, col_point)

            for _, _v in ret_dict[tune].items():
                self.set_cell_style(sheet, row_point, col_point, _v.strip(" "),
                                    self.alignment_center, self.color_data, self.font_data)
                row_point += 1

        return True

class Speccpu2006(BenchMark):
    def __init__(self):
        super().__init__()
        self.tool_name = TOOL_NAME_CPU06
        self.dtype = ["fp", "int"]
        self.tune = ["base", "peak"]
        self.items = {
            self.dtype[0]: ["410.bwaves", "416.gamess",
                            "433.milc", "434.zeusmp",
                            "435.gromacs", "436.cactusADM",
                            "437.leslie3d", "444.namd",
                            "447.dealII", "450.soplex",
                            "453.povray", "454.calculix",
                            "459.GemsFDTD", "465.tonto",
                            "470.lbm", "481.wrf",
                            "482.sphinx3", "SPECfp_2006"],
            self.dtype[1]: ["400.perlbench", "401.bzip2",
                            "403.gcc", "429.mcf", "445.gobmk",
                            "456.hmmer", "458.sjeng",
                            "462.libquantum", "464.h264ref",
                            "471.omnetpp", "473.astar",
                            "483.xalancbmk", "SPECint_2006"]}
        self.thread = ["单线程", "多线程"]
        self.cols_width = [10, 10, self.ret_col_2_width]

    def set_row_header(self, sheet: Worksheet, row_point: int, data: dict):
        if row_point > sheet.max_row:
            self.set_cell_style(sheet, row_point, self.col_point_start, data["thread"], self.alignment_center, self.color_item_1, self.font_item_1)
            self.set_cell_style(sheet, row_point, self.col_point_start + 1, data["type"], self.alignment_center, self.color_item_1, self.font_item_1)
            for i, _l in enumerate(self.items[data["type"]]):
                self.set_cell_style(sheet, row_point + i, self.col_point_start + 2, _l, self.alignment_left, self.color_item_2, self.font_item_2)

            self.merge_col_cell_by_value(sheet, utils.get_column_letter(self.col_point_start + 1), row_point, row_point + len(self.items[data["type"]]) - 1, data["type"])
            self.merge_col_cell_by_value(sheet, utils.get_column_letter(self.col_point_start), row_point, row_point + len(self.items[data["type"]]) - 1, data["thread"])
        pass

    def ret_to_dict(self, file: str):
        ret_dict = {"tool_name": self.tool_name, "items": {}}
        ret_names_json = None
        with open(file, 'r') as f:
            try:
                ret_names_json = json.loads(f.read())
            except Exception as e:
                print("json 文件打开失败!", e)
                return None
        for _file in list(ret_names_json.values()):
            if len(_file) == 0:
                continue
            _file_lines = _file.split("\n")
            _thread = None
            _read_flag = False
            _type = None
            _items_list = []
            _scores_tune = {self.tune[0]: [], self.tune[1]: []}
            for i, line in enumerate(_file_lines):
                if "========" in line:
                    _start_key = _file_lines[i+1].strip().split(" ")[0]
                    if _start_key in self.items[self.dtype[0]]:
                        _type = self.dtype[0]
                        _items_list = self.items[_type]
                    elif _start_key in self.items[self.dtype[1]]:
                        _type = self.dtype[1]
                        _items_list = self.items[_type]
                    else:
                        return None
                    _read_flag = True
                    continue
                if _read_flag:
                    _line = line
                    if _line.count("*") == 1:
                        if _line.index("*") > 50:
                            _line = _line.split("*")
                            _line.insert(0, " ")
                        else:
                            _line = _line.split("*")
                    elif _line.count("*") == 2:
                        _line = _line.split("*")
                    elif "SPEC" in _line and "base2006" in _line:
                        _line = [line, _file_lines[i+1]]
                        _read_flag = False
                    else:
                        _line = ["NR/RE", "NR/RE"]
                    for j, _l in enumerate(_line[:2]):
                        temp = re.findall(r" +\d+\.?\d*e?-?\d*?", _l)
                        _str = "NR/RE"
                        if len(temp) > 0:
                            _str = temp[-1].strip()
                            if _thread is None:
                                _thread = (self.thread[1] if temp[0].strip() > '1' else self.thread[0])
                        _scores_tune[self.tune[j]].append(_str)
                    if not _read_flag:
                        break
            if not _thread or not _type:
                print("线程数为%s,测试类型为%s!" % (_thread, _type))
                continue
            _items_key = _thread+"_"+_type
            score_dict = {}
            for k, v in _scores_tune.items():
                score_dict[k] = dict(zip(_items_list, v))
            ret_dict["items"][_items_key] = score_dict
        return ret_dict

    def ret_dict_to_excel(self, workbook: Workbook, ret_dict: dict):
        for k, v in ret_dict["items"].items():
            for _k, _v in v.items():
                sheet_name = "%s(%s)" % (self.tool_name, _k)
                sheet = None
                if sheet_name in workbook.sheetnames:  # 已存在sheet
                    sheet = workbook[sheet_name]
                else:   # 不存在sheet
                    sheet = workbook.create_sheet(sheet_name)
                    self.set_col_items(sheet)

                _thread = k.split("_")[0]
                _dtype = k.split("_")[1]

                row_point = self.find_row_point(sheet, {"thread": _thread, "type": _dtype})
                self.set_row_header(sheet, row_point, {"thread": _thread, "type": _dtype})

                col_point = self.find_col_point(sheet, row_point)
                self.set_col_title(sheet, col_point)

                for i, (_, __v) in enumerate(_v.items()):
                    self.set_cell_style(sheet, row_point + i, col_point, __v,
                                        self.alignment_center, self.color_data, self.font_data)

        return True


class Speccpu2017(BenchMark):
    def __init__(self):
        super().__init__()
        self.tool_name = TOOL_NAME_CPU17
        self.dtype = ["fp", "int"]
        self.items = {
            self.dtype[0]: {
                "rate": [
                    "503.bwaves_r", "507.cactuBSSN_r",
                    "508.namd_r", "510.parest_r",
                    "511.povray_r", "519.lbm_r",
                    "521.wrf_r", "526.blender_r",
                    "527.cam4_r", "538.imagick_r",
                    "544.nab_r", "549.fotonik3d_r",
                    "554.roms_r",
                    "SPECrate2017_fp"],
                "speed": [
                    "603.bwaves_s",
                    "607.cactuBSSN_s",
                    "619.lbm_s", "621.wrf_s",
                    "627.cam4_s", "628.pop2_s",
                    "638.imagick_s", "644.nab_s",
                    "649.fotonik3d_s", "654.roms_s",
                    "SPECspeed2017_fp"]},
            self.dtype[1]: {
                "rate": [
                    "500.perlbench_r", "502.gcc_r",
                    "505.mcf_r", "520.omnetpp_r",
                    "523.xalancbmk_r", "525.x264_r",
                    "531.deepsjeng_r", "541.leela_r",
                    "548.exchange2_r", "557.xz_r",
                    "SPECrate2017_int"],
                "speed": [
                    "600.perlbench_s", "602.gcc_s",
                    "605.mcf_s", "620.omnetpp_s",
                    "623.xalancbmk_s",
                    "625.x264_s",
                    "631.deepsjeng_s",
                    "641.leela_s",
                    "648.exchange2_s",
                    "657.xz_s",
                    "SPECspeed2017_int"]}}
        self.thread = ["单线程", "多线程"]
        self.tune = ["base", "peak"]
        self.cols_width = [10, 10, 10, self.ret_col_2_width]

    def set_row_header(self, sheet: Worksheet, row_point: int, data: dict):
        if row_point > sheet.max_row:
            self.set_cell_style(sheet, row_point, self.col_point_start, data["thread"], self.alignment_center, self.color_item_1, self.font_item_1)
            self.set_cell_style(sheet, row_point, self.col_point_start + 1, data["type"], self.alignment_center, self.color_item_1, self.font_item_1)
            self.set_cell_style(sheet, row_point, self.col_point_start + 2, data["tune"], self.alignment_center, self.color_item_1, self.font_item_1)
            for i, _l in enumerate(self.items[data["type"]][data["tune"]]):
                self.set_cell_style(sheet, row_point + i, self.col_point_start + 3, _l, self.alignment_left, self.color_item_2, self.font_item_2)

            self.merge_col_cell_by_value(sheet, utils.get_column_letter(self.col_point_start + 2), row_point, row_point + len(self.items[data["type"]][data["tune"]]) - 1, data["tune"])
            self.merge_col_cell_by_value(sheet, utils.get_column_letter(self.col_point_start + 1), row_point, row_point + len(self.items[data["type"]][data["tune"]]) - 1, data["type"])
            self.merge_col_cell_by_value(sheet, utils.get_column_letter(self.col_point_start), row_point, row_point + len(self.items[data["type"]][data["tune"]]) - 1, data["thread"])
        pass

    def ret_to_dict(self, file: str):
        ret_dict = {"tool_name": self.tool_name, "items": {}}
        ret_names_json = None
        with open(file, 'r') as f:
            try:
                ret_names_json = json.loads(f.read())
            except Exception as e:
                print("json 文件打开失败!", e)
                return None

        for _file in list(ret_names_json.values()):
            if not _file:
                continue
            _thread = None
            _read_flag = False
            _type_int_fp = None
            _type_rate_speed = None
            _file_lines = _file.split("\n")
            _scores_tune = []
            _scores_tune = {self.tune[0]: [], self.tune[1]: []}
            for i, line in enumerate(_file_lines):
                if "========" in line:
                    _read_flag = True
                    continue
                if _read_flag:
                    _line = line
                    if _line.count("*") == 1:
                        if _line.index("*") > 50:
                            _line = _line.split("*")
                            _line.insert(0, " ")
                        else:
                            _line = _line.split("*")
                    elif _line.count("*") == 2:
                        _line = _line.split("*")
                    elif all(key in line for key in [
                            "2017", "base"]):
                        _regx = r"(?<=SPEC)[a-z]*|(?<=_)[a-z]*"
                        _flag = re.findall(_regx, _line)
                        _type_rate_speed = _flag[0]
                        _type_int_fp = _flag[1]
                        _line = [line, _file_lines[i+1]]
                        _read_flag = False
                    else:
                        _line = ["NR/RE", "NR/RE"]
                    for j, _l in enumerate(_line[:2]):
                        temp = re.findall(r" +\d+\.?\d*e?-?\d*?", _l)
                        _str = ''
                        if len(temp) > 0:
                            _str = temp[-1].strip()
                            if not _thread:
                                _thread = (self.thread[1] if temp[0].strip() > '1' else self.thread[0])
                        else:
                            _str = "NR/RE"
                        _scores_tune[self.tune[j]].append(_str)
                    if not _read_flag:
                        break
            _score_key = "%s_%s_%s" % (_thread, _type_int_fp, _type_rate_speed)
            _items_list = self.items[_type_int_fp][_type_rate_speed]
            score_dict = {}
            for k, v in _scores_tune.items():
                score_dict[k] = dict(zip(_items_list, v))
            ret_dict["items"][_score_key] = score_dict
        return ret_dict

    def ret_dict_to_excel(self, workbook: Workbook, ret_dict: dict):
        for k, v in ret_dict["items"].items():
            for _k, _v in v.items():
                sheet_name = "%s(%s)" % (self.tool_name, _k)
                sheet = None
                if sheet_name in workbook.sheetnames:  # 已存在sheet
                    sheet = workbook[sheet_name]
                else:   # 不存在sheet
                    sheet = workbook.create_sheet(sheet_name)
                    self.set_col_items(sheet)

                _thread = k.split("_")[0]
                _dtype = k.split("_")[1]
                _tune = k.split("_")[2]

                row_point = self.find_row_point(sheet, {"thread": _thread, "type": _dtype, "tune": _tune})
                self.set_row_header(sheet, row_point, {"thread": _thread, "type": _dtype, "tune": _tune})

                col_point = self.find_col_point(sheet, row_point)
                self.set_col_title(sheet, col_point)

                for i, (_, __v) in enumerate(_v.items()):
                    self.set_cell_style(sheet, row_point + i, col_point, __v,
                                        self.alignment_center, self.color_data, self.font_data)

        return True


class Iozone(BenchMark):
    def __init__(self):
        super().__init__()
        self.tool_name = TOOL_NAME_IOZONE
        self.items = {
            "block_size": "",
            "rw_items": ["写测试（KB/s）", "重写测试（KB/s）", "读测试（KB/s）",
                         "重读测试（KB/s）", "随机读测试（KB/s）", "随机写测试（KB/s）"]}
        self.cols_width = [self.ret_col_1_width/2, self.ret_col_2_width/2]

    def set_row_header(self, sheet: Worksheet, row_point: int, record):
        if row_point > sheet.max_row:
            for idx in range(0, len(self.items['rw_items'])):
                self.set_cell_style(sheet, row_point + idx, self.col_point_start, self.items['rw_items'][idx], self.alignment_center, self.color_item_1, self.font_item_1)
                self.set_cell_style(sheet, row_point + idx, self.col_point_start + 1, record['文件大小'], self.alignment_center, self.color_item_2, self.font_item_2)
                sheet.row_dimensions[row_point + idx].height = 30
            self.merge_col_cell(sheet, utils.get_column_letter(self.col_point_start + 1), row_point, row_point + len(self.items['rw_items']) - 1)
        pass

    def find_row_point(self, sheet: Worksheet, record):
        if sheet.max_row >= self.row_data_start:
            for row in sheet.iter_rows(min_row = self.row_data_start):
                if row[self.col_point_start - 1 + 1].value == record['文件大小']:
                    return row[self.col_point_start - 1].row
        return sheet.max_row + 1

    def ret_to_dict(self, file: str):
        ret_dict = {"tool_name": self.tool_name}
        with open(file, 'r') as f:
            lines = f.readlines()
        ret_dict['测试记录'] = []
        find_result = False
        for line in lines:
            if find_result == False:
                if line.strip()[0:2] == "kB":
                    find_result = True
            else:
                record = line.strip().split(' ')
                while '' in record:
                    record.remove('')
                ret_dict['测试记录'].append({
                    '文件大小': int(record[0]),
                    '块大小': int(record[1]),
                    '写测试（KB/s）': int(record[2]),
                    '重写测试（KB/s）': int(record[3]),
                    '读测试（KB/s）': int(record[4]),
                    '重读测试（KB/s）': int(record[5]),
                    '随机读测试（KB/s）': int(record[6]),
                    '随机写测试（KB/s）': int(record[7])})
                break
        return ret_dict

    def ret_dict_to_excel(self, workbook: Workbook, ret_dict: dict):
        if ret_dict['tool_name'] in workbook.sheetnames:
            sheet = workbook[ret_dict['tool_name']]
        else:
            sheet = workbook.create_sheet(ret_dict['tool_name'])
            self.set_col_items(sheet)

        for record in ret_dict['测试记录']:
            row_point = self.find_row_point(sheet, record)
            self.set_row_header(sheet, row_point, record)

            col_point = self.find_col_point(sheet, row_point)
            self.set_col_title(sheet, col_point)

            for idx in range(0, len(self.items['rw_items'])):
                self.set_cell_style(sheet, row_point + idx, col_point, record[self.items['rw_items'][idx]],
                            self.alignment_center, self.color_data, self.font_data)

        return True


class Specjvm2008(BenchMark):
    def __init__(self):
        super().__init__()
        self.tool_name = TOOL_NAME_JVM08
        self.items = ["compiler", "compress", "crypto",
                      "derby", "mpegaudio", "scimark.large",
                      "scimark.small", "serial", "startup",
                      "sunflow", "xml",
                      "Noncompliant composite result"]
        self.tune = ["base", "peak"]
        self.cols_width = [10, self.ret_col_2_width]

    def set_row_header(self, sheet: Worksheet, row_point: int, data: dict):
        if row_point > sheet.max_row:
            self.set_cell_style(sheet, row_point, self.col_point_start, data["tune"], self.alignment_center, self.color_item_1, self.font_item_1)
            for i, _l in enumerate(self.items):
                self.set_cell_style(sheet, row_point + i, self.col_point_start + 1, _l, self.alignment_center, self.color_item_2, self.font_item_2)
            self.merge_col_cell(sheet, utils.get_column_letter(self.col_point_start), row_point, row_point + len(self.items) - 1)
        pass

    def ret_to_dict(self, file: str):
        ret_dict = {"tool_name": self.tool_name, "items": {}}
        _read_flag = False
        _items_list = self.items
        _tune = None
        _file_lines = []
        with open(file, 'r') as f:
            _file_lines = f.readlines()
        _score_list = []
        _score_dict = {}
        for line in _file_lines:
            _tmp = re.findall(r"base|peak", line, re.IGNORECASE)
            if _tmp:
                _tune = _tmp[0].lower()
                _read_flag = True
                continue
            if _read_flag:
                if self.items[-1] in line:
                    _items_list[-1] = _items_list[-1]+":"
                    _read_flag = False
                regx = "|".join("(?<=%s)" % s for s in _items_list)
                regx = r"(%s)\s+[a-zA-Z0-9.]*" % regx
                temp = re.search(regx, line, re.IGNORECASE)
                if not temp:
                    continue
                _score_list.append(temp.group().strip())
                if not _read_flag:
                    break
        _score_dict = dict(zip(self.items, _score_list))
        ret_dict["items"][_tune] = _score_dict
        return ret_dict

    def ret_dict_to_excel(self, workbook: Workbook, ret_dict: dict):
        sheet_name = "%s" % (self.tool_name)
        if sheet_name in workbook.sheetnames:  # 已存在sheet
            sheet = workbook[sheet_name]
        else:   # 不存在sheet
            sheet = workbook.create_sheet(sheet_name)
            self.set_col_items(sheet)

        for k, v in ret_dict["items"].items():
            row_point = self.find_row_point(sheet, {"tune": k})
            self.set_row_header(sheet, row_point, {"tune": k})

            col_point = self.find_col_point(sheet, row_point)
            self.set_col_title(sheet, col_point)

            for _, _v in v.items():
                self.set_cell_style(sheet, row_point, col_point, _v,
                                    self.alignment_center, self.color_data, self.font_data)
                row_point += 1

        return True


class Lmbench(BenchMark):
    def __init__(self):
        super().__init__()
        self.tool_name = TOOL_NAME_LMB
        self.items = {
            "Basic system parameters": ["Mhz", "tlb pages",
                                        "cache line bytes",
                                        "mem par", "scal load"],
            "Processor": ["Mhz", "null call", "null I/O", "stat",
                          "open close", "slct TCP", "sig inst",
                          "sig hndl", "fork proc", "exec proc",
                          "sh proc"],
            "Basic integer operations": ["intgr bit", "intgr add",
                                         "intgr mul", "intgr div",
                                         "intgr mod"],
            "Basic uint64 operations": ["int64 bit", "int64 add",
                                        "int64 mul", "int64 div",
                                        "int64 mod"],
            "Basic float operations": ["float add", "float mul",
                                       "float div", "float bogo"],
            "Basic double operations": ["double add", "double mul",
                                        "double div", "double bogo"],
            "Context switching": ["2p/0K", "2p/16K",
                                  "2p/64K", "8p/16K",
                                            "8p/64K", "16p/16K",
                                            "16p/64K"],
            "*Local* Communication latencies": ["2p/0K", "Pipe",
                                                "AF UNIX", "UDP",
                                                "RPC/UDP", "TCP",
                                                "RPC/TCP", "TCP conn"
                                                ],
            "File & VM system latencies in microseconds":
            ["0K File create", "0K File delete",
             "10K File create", "10K File delete",
             "Mmap Latency", "Prot Fault", "Page Fault",
             "100fd selct",
             ],
            "*Local* Communication bandwidths in MB/s - bigger is better":
            ["Pipe", "AF UNIX", "TCP", "File reread", "Mmap reread",
             "Bcopy(libc)", "Bcopy(hand)", "Mem read", "Mem write"],
            "Memory latencies in nanoseconds": ["Mhz", "L1 $",
                                                "L2 $", "Main mem",
                                                "Rand mem"]
        }
        self.cols_width = [self.ret_col_1_width, self.ret_col_2_width]

    def set_row_header(self, sheet: Worksheet, row_point: int, ret_dict):
        if row_point > sheet.max_row:
            # 填充第一列
            for k, v in self.items.items():
                self.set_cell_style(sheet, row_point, self.col_point_start, k,
                                    self.alignment_center, self.color_item_1, self.font_item_1)
                if len(v) > 1:
                    # 填充第二列
                    for i, s in enumerate(v):
                        self.set_cell_style(sheet, row_point + i, self.col_point_start + 1, s,
                                            self.alignment_center, self.color_item_2, self.font_item_2)
                    self.merge_col_cell(sheet, utils.get_column_letter(self.col_point_start), row_point, row_point + len(v) - 1)
                    row_point += len(v) - 1
                row_point += 1
        pass

    def find_row_point(self, sheet: Worksheet, data: dict):
        return self.row_data_start

    def parse_target_len(self, t_line):
        target_lens = list()
        if not t_line:
            return target_lens

        raw_target_lens = [len(x) for x in t_line.strip("\n").split(' ')]
        offset = 0
        for i in raw_target_lens:
            if i == 0:
                offset += 1
                continue
            target_lens.append(offset + i)
            offset = 0
        return target_lens

    def parse_data(self, data, target_lens):
        result_data = list()
        if not data or not target_lens:
            return result_data

        start = 0
        for i in target_lens:

            if start == 0:
                result = str(data[start: start + i])
                start += i
            else:
                result = str(data[start: start + i + 1])
                start += i + 1
            result = result.strip('\n .')
            if result.endswith("K") or result.endswith("k"):
                temp = re.search(r'\d+\.?\d*', result).group()
                result = str(float(temp) * 1000)
            result_data.append(result)
        return result_data

    def ret_to_dict(self, file):
        ret_dict = {"tool_name": self.tool_name}
        with open(file, 'r') as f:
            file_lines = f.readlines()
        _scores = []
        n = -2000
        for line in file_lines:
            if n < 0:
                for keywords in self.items.keys():
                    if keywords in line:
                        if keywords == "Basic system parameters":
                            n = 0
                        else:
                            n = 1
                        break
            if n == 5:
                target_line = "" + line  # type: str
                _scores_temp = list()
            if n >= 6 and not line.isspace() and 'make' not in line:
                data_line = line
                _score_temp = self.parse_data(
                    data_line, self.parse_target_len(target_line))
                if keywords == "Basic system parameters":
                    _scores_temp.append(_score_temp[3:])
                else:
                    _scores_temp.append(_score_temp[2:])
            if n >= 6 and (line.isspace() or line == file_lines[-1]):
                test_num = len(_scores_temp)
                if len(_scores) == test_num:
                    for i, v in enumerate(_scores_temp):
                        _scores[i] += v
                        # _scores[i].extend[v]
                else:
                    for i, v in enumerate(_scores_temp):
                        _scores.append([])
                        _scores[i] = v
                n = -2000
            n += 1
        ret_dict["items"] = []
        for _score in _scores:
            mid_ret_dict = []
            for (key, value) in self.items.items():
                mid_ret_dict.append(
                    {key: dict(zip(value, _score[:len(value)]))})
                del _score[:len(value)]
            ret_dict["items"].append(mid_ret_dict)
        return ret_dict

    def ret_dict_to_excel(self, workbook: Workbook, ret_dict: dict):
        sheet = None
        if self.tool_name in workbook.sheetnames:  # 已存在sheet
            sheet = workbook[self.tool_name]
        else:   # 不存在sheet
            sheet = workbook.create_sheet(self.tool_name)
            self.set_col_items(sheet)

        row_point = self.find_row_point(sheet, ret_dict)
        self.set_row_header(sheet, row_point, ret_dict)

        col_point = self.find_col_point(sheet, row_point)
        self.set_col_title(sheet, col_point)

        for i, item_list in enumerate(ret_dict["items"]):
            j_index = row_point
            for item in item_list:
                _items = list(item.values())[0]
                for _ in _items.keys():
                    cell_key = sheet.cell(j_index, self.col_point_start - 1 + 2).value
                    self.set_cell_style(sheet, j_index, col_point, _items.get(cell_key),
                                        self.alignment_center, self.color_data, self.font_data)
                    j_index += 1

        return True


class Stream(BenchMark):
    def __init__(self):
        super().__init__()
        self.tool_name = TOOL_NAME_STREAM
        self.items = ["Array size", "Copy", "Scale", "Add", "Triad"]
        self.thread = ["单线程", "多线程"]
        self.cols_width = [10, self.ret_col_2_width]

    def set_row_header(self, sheet: Worksheet, row_point: int, data: dict):
        if row_point > sheet.max_row:
            self.set_cell_style(sheet, row_point, self.col_point_start, data["tune"], self.alignment_center, self.color_item_1, self.font_item_1)
            for i, _l in enumerate(self.items):
                self.set_cell_style(sheet, row_point + i, self.col_point_start + 1, _l, self.alignment_center, self.color_item_2, self.font_item_2)
            self.merge_col_cell(sheet, utils.get_column_letter(self.col_point_start), row_point, row_point + len(self.items) - 1)
        pass

    def ret_to_dict(self, file: str):
        stream_result_str = None
        _single_result = None
        _multiple_result = None

        ret_dict = {"tool_name": self.tool_name}

        score_list = ['' for i in self.items]
        ret_dict[self.thread[0]] = dict(
            zip(self.items, score_list))
        ret_dict[self.thread[1]] = dict(
            zip(self.items, score_list))

        # 打印字典中的内容
        # print(ret_dict)
        # 读取单核测试结果和多核测试结果
        with open(file, 'r') as f:
            stream_result_str = f.read()
        if stream_result_str:
            stream_json = json.loads(stream_result_str)
            # print(stream_json['single'])
            # print(stream_json['multiple'])
            _single_result = stream_json['single']
            _multiple_result = stream_json['multiple']

            # 处理单核测试结果
            if (_single_result is not None and len(_single_result) > 0):
                for line in _single_result.split('\n'):
                    for k in self.items:
                        if k in line:
                            temp = re.findall(r"-?\d+\.?\d*e?-?\d*?", line)
                            if temp[0]:
                                ret_dict[self.thread[0]][k] = temp[0]
            else:
                print("stream 测试没有单核测试结果")

            # 处理多核测试结果
            if (_multiple_result is not None and len(_multiple_result) > 0):
                for line in _multiple_result.split('\n'):
                    for k in self.items:
                        if k in line:
                            temp = re.findall(r"-?\d+\.?\d*e?-?\d*?", line)
                            if temp[0]:
                                ret_dict[self.thread[1]][k] = temp[0]
            else:
                print("stream 测试没有多核测试结果")

            # print(ret_dict)
        if len(ret_dict.keys()) > 1:
            return ret_dict
        else:
            return None

    def ret_dict_to_excel(self, workbook: Workbook, ret_dict: dict):
        sheet = None
        if self.tool_name in workbook.sheetnames:  # 已存在sheet
            sheet = workbook[self.tool_name]
        else:   # 不存在sheet
            sheet = workbook.create_sheet(self.tool_name)
            self.set_col_items(sheet)

        for tune in self.thread:
            if tune not in ret_dict.keys():
                continue
            row_point = self.find_row_point(sheet, {"tune": tune})
            self.set_row_header(sheet, row_point, {"tune": tune})

            col_point = self.find_col_point(sheet, row_point)
            self.set_col_title(sheet, col_point)

            for _, _v in ret_dict[tune].items():
                self.set_cell_style(sheet, row_point, col_point, _v.strip(" "),
                                    self.alignment_center, self.color_data, self.font_data)
                row_point += 1

        return True


class Netperf(BenchMark):
    def __init__(self):
        super().__init__()
        self.tool_name = TOOL_NAME_NETPERF
        self.items = ["TCP_Stream", "UDP_Stream", "网络响应时间",
                      "TCP_CRR", "TCP_RR", "UDP_RR"]
        self.cols_width = [self.ret_col_2_width]

    def set_row_header(self, sheet: Worksheet, row_point: int, data: dict = None):
        if row_point > sheet.max_row:
            for i, _l in enumerate(self.items):
                self.set_cell_style(sheet, row_point + i, self.col_point_start, _l, self.alignment_center, self.color_item_1, self.font_item_1)
        pass

    def find_row_point(self, sheet: Worksheet, data: dict = None):
        return self.row_data_start

    def ret_to_dict(self, file: str):
        ret_dict = {"tool_name": self.tool_name}
        ret_dict['测试记录'] = []
        with open(file, 'r') as f:
            lines = f.readlines()
            for i in [6, 13, 22, 28, 36, 44]:
                ret_dict['测试记录'].append(lines[i].strip().split(' ')[-1])
        return ret_dict

    def ret_dict_to_excel(self, workbook: Workbook, ret_dict: dict):
        print(ret_dict)
        sheet = None
        if self.tool_name in workbook.sheetnames:
            sheet = workbook[self.tool_name]
        else:
            sheet = workbook.create_sheet(self.tool_name)
            self.set_col_items(sheet)
        
        row_point = self.find_row_point(sheet)
        self.set_row_header(sheet, row_point)

        col_point = self.find_col_point(sheet, row_point)
        self.set_col_title(sheet, col_point)

        for _v in ret_dict['测试记录']:
            self.set_cell_style(sheet, row_point, col_point, _v.strip(" "),
                                self.alignment_center, self.color_data, self.font_data)
            row_point += 1

        return True


class Fio(BenchMark):
    def __init__(self):
        super().__init__()
        self.tool_name = TOOL_NAME_FIO
        self.items = {"items": ["bs", "io", "iops", "bw"]}
        self.cols_width = [self.ret_col_1_width/2, self.ret_col_2_width/2]

    def set_row_header(self, sheet: Worksheet, row_point: int, ret_dict):
        if row_point > sheet.max_row:
            self.set_cell_style(sheet, row_point, self.col_point_start, ret_dict['rw'], self.alignment_center, self.color_item_1, self.font_item_1)
            for i, _l in enumerate(self.items["items"]):
                self.set_cell_style(sheet, row_point + i, self.col_point_start + 1, _l, self.alignment_center, self.color_item_2, self.font_item_2)
            self.merge_col_cell(sheet, utils.get_column_letter(self.col_point_start), row_point, row_point + len(self.items["items"]) - 1)
        pass

    def find_row_point(self, sheet: Worksheet, ret_dict):
        if sheet.max_row >= self.row_data_start:
            for row in sheet.iter_rows(min_row = self.row_data_start):
                if row[self.col_point_start - 1].value == ret_dict['rw']:
                    for rv in row[self.col_point_start - 1 + 2:]:
                        if rv.value == ret_dict["items"]["bs"]:
                            return row[self.col_point_start - 1].row
        return sheet.max_row + 1
 
    def ret_to_dict(self, file: str):
        ret_dict = {"tool_name": self.tool_name}
        units = {"k":1e3, "M":1e6, "G":1e9}
        _file_lines = None
        _rw = None
        with open(file, "r") as f:
            _file_lines = f.readlines()
        if not _file_lines:
            print("file is empty")
            return None
        _regx = r'(?<=rw=).[a-z_]*'
        _rw = re.findall(_regx, _file_lines[0])
        if len(_rw) > 0:
            _rw = _rw[0]
        else:
            print("Fio:rw not found")
            return None
        _regx = r'(?<=bs=\(R\)\s).[A-Z0-9_\.]*'
        _bs = re.findall(_regx, _file_lines[0])
        if len(_bs) > 0:
            _bs = _bs[0]
        else:
            print("Fio:bs not found")
            return None
        _read_flag = "read"
        _read_flag = _read_flag if _read_flag in _rw else "write"
        _flags_list = [_read_flag, "IOPS", "BW"]
        _iops = 0
        _bw = 0
        _io = 0
        for line in _file_lines[1:]:
            line = line.strip(" ")
            if all(_f in line for _f in _flags_list):
                _regx = r'(?<=IOPS=).*,'
                _iops = re.findall(_regx, line)[0][:-1]
                for unit in units:
                    if unit in _iops:
                        _iops = float(_iops.replace(unit, '')) * units[unit]
                        break
                _regx = r'(?<=BW=).*\)'
                _bw_io = re.findall(_regx, line)
                _bw_list = _bw_io[0].split(")")[0].split("(")
                _bw = _bw_list[0].strip(" ") + " (" + _bw_list[1] + ")"
                _io = _bw_io[0].split(")")[1].split("(")[1].split("/")[0]
                break
        ret_dict["rw"] = _rw
        ret_dict["items"] = dict(
            zip(self.items["items"], [_bs, _io, _iops, _bw]))
        return ret_dict

    def ret_dict_to_excel(self, workbook: Workbook, ret_dict: dict):
        sheet = None
        if self.tool_name in workbook.sheetnames:  # 已存在sheet
            sheet = workbook[self.tool_name]
        else:   # 不存在sheet
            sheet = workbook.create_sheet(self.tool_name)
            self.set_col_items(sheet)

        row_point = self.find_row_point(sheet, ret_dict)
        self.set_row_header(sheet, row_point, ret_dict)

        col_point = self.find_col_point(sheet, row_point)
        self.set_col_title(sheet, col_point)

        for i, v in enumerate(ret_dict["items"].values()):
            self.set_cell_style(sheet, row_point + i, col_point, v,
                                self.alignment_center, self.color_data, self.font_data)

        return True


class ExportXlsx(object):
    def __init__(self):
        self.env_dict = None
        self.ret_dict = None
        self.tool_name = None
        self.tool_cls = None

        self.excel_name = "kytuning-result.xlsx"
        self.jsob_name = "kytuning-result.json"

    @ staticmethod
    def get_files(ret_path: str) -> list:
        files = []
        if os.path.isfile(ret_path):
            return [ret_path]
        try:
            files = [os.path.join(ret_path, file)
                     for file in os.listdir(ret_path)
                     if os.path.isfile(os.path.join(ret_path, file))]
        except Exception as e:
            print("列出文件失败!", e)
        return files

    @ staticmethod
    def get_tool_cls(tool_name):
        """通过工具名实例化使用的类"""
        cls_dict = {v.__name__: v for v in BenchMark.__subclasses__()}
        match_names = difflib.get_close_matches(
            tool_name, list(cls_dict.keys()), 1)
        return cls_dict.get(match_names[0])()

    def ret_to_dict(
            self, tool_name: str, ret_path: str, cmd: str, argv: str) -> list:
        """结果解析转化为字典"""
        ret_dict = {}
        _tool_name = tool_name
        _ret_path = ret_path
        self.tool_cls = self.get_tool_cls(_tool_name)
        if self.tool_cls is None:
            print("未能实例化工具类!\n")
            return ret_dict
        self.tool_cls.value_cmd = cmd
        self.tool_cls.value_modify_args = argv
        self.tool_cls.tool_name = _tool_name
        return self.tool_cls.ret_to_dict(_ret_path)

    def rets_to_dict_list(
            self, tool_name: str, ret_path: str, cmd: str, argv: str) -> list:
        ret_dict_list = []
        for file in self.get_files(ret_path):
            ret_dict = self.ret_to_dict(tool_name, file, cmd , argv)
            if ret_dict is None:
                continue
            ret_dict_list.append(ret_dict)
        return ret_dict_list

    def write_to_xlsx(
            self, ret_dict: dict, excel_path: str):
        if not ret_dict:
            print("无结果可转化!")
            return False
        excel_file = excel_path + "/" + self.excel_name
        _ret_dict = ret_dict
        wb = None
        if os.path.isfile(excel_file):
            wb = load_workbook(excel_file)
        else:
            wb = Workbook()
            wb.remove(wb.active)
        if not self.tool_cls.ret_dict_to_excel(wb, _ret_dict):
            print("结果字典转excel失败!")
            return False
        wb.save(excel_file)
        return True

    def export_env_to_xlsx(self, env_dict: dict, excel_path: str):
        wb = None
        sheet = None
        excel_file = excel_path + "/" + self.excel_name
        env_write = BenchMark()
        if os.path.isfile(excel_file):
            wb = load_workbook(excel_file)
            if env_write.sheet_env_title in wb.sheetnames:
                sheet = wb[env_write.sheet_env_title]
        else:
            wb = Workbook()
            sheet = wb.active
            sheet.title = env_write.sheet_env_title
        env_write.env_dict_to_excel(sheet, env_dict)
        wb.save(excel_file)

    def export_ret_to_xlsx(
            self, tool_name: str, ret_path: str,
            export_path: str, cmd: str, argv: str) -> bool:
        _tool_name = tool_name.title()
        _ret_path = ret_path
        _ret_dict_list = []
        if not _tool_name or not _ret_path:
            return False
        if not os.path.exists(_ret_path):
            print("%s结果文件（路径）不存在!" % (ret_path))
            return False

        if os.path.isfile(_ret_path):
            _ret_dict_list.append(self.ret_to_dict(
                _tool_name, _ret_path, cmd, argv))
        else:
            _ret_dict_list = self.rets_to_dict_list(
                _tool_name, _ret_path, cmd, argv)
        if not _ret_dict_list:
            print("获取结果字典列表为空!")
            return False
        for _ret_dict in _ret_dict_list:
            self.write_to_xlsx(_ret_dict, export_path)
        return True


support_tools = {'lmbench': TOOL_NAME_LMB, 'unixbench': TOOL_NAME_UB, 'speccpu2006': TOOL_NAME_CPU06,
                 'speccpu2017': TOOL_NAME_CPU17, 'iozone': TOOL_NAME_IOZONE, 'specjvm2008': TOOL_NAME_JVM08,
                 'stream': TOOL_NAME_STREAM, 'netperf': TOOL_NAME_NETPERF, 'fio': TOOL_NAME_FIO}


def constructParse():

    parser = argparse.ArgumentParser(
        description="Export benchmark raw results to excel format.")

    parser.add_argument('-n', '--name', choices=support_tools.keys(), help='benchmark name.',
                        type=str)

    parser.add_argument('-f', '--file', help='absolute path of benchmark raw result.',
                        type=str)

    parser.add_argument('-o', '--output', help='the dir where we export excel to.',
                        default=os.getcwd(), type=str)
    return parser


if __name__ == '__main__':

    parse = constructParse()

    try:
        args = parse.parse_args()
        if args is not None:
            tool_name = support_tools[args.name]
            xlsx_obj = ExportXlsx()

            path = args.output
            xlsx_obj.export_ret_to_xlsx(
                tool_name, args.file,
                path, 'cmd', 'args')
    except Exception as e:
        parse.print_help()
